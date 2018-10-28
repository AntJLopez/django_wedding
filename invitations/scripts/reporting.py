from openpyxl import load_workbook
from guests.models import Guest


def mark_invites_sent(filename='invitations/records/invites_printed.xlsx'):
    wb = load_workbook(filename)
    ws = wb.active

    for userfield in ws['B']:
        username, row = userfield.value, userfield.row
        if row == 1:
            continue
        guest = Guest.objects.get(username=username)
        user_id = guest.id
        ws[f'C{row}'] = user_id
        print(guest.username)
        print(user_id)
        guest.invite_printed = True
        guest.save()

    wb.save(filename='invitations/records/output.xlsx')


def run():
    mark_invites_sent()
